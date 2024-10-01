import os
import openpyxl as ex
way =  os.getcwd() 
print(way)

wb = ex.load_workbook('book.xlsx')
sheet_names = wb.sheetnames 
print(sheet_names)
sheet = wb['Sheet1']
print(sheet.cell( row = 2, column = 2 ).value )
sheet.cell( row = 2, column = 2, value = 'potato' )
wb.save('book3.xlsx')